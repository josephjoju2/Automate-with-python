import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
print(get_column_letter(27))
print(get_column_letter(3))
print(column_index_from_string('AB'))
wb=openpyxl.load_workbook('example.xlsx')
sheet=wb.active
print(get_column_letter(sheet.max_column))

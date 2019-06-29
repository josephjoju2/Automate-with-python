import openpyxl
wb=openpyxl.load_workbook('example.xlsx')
sheet=wb.get_sheet_by_name('Sheet1')
print(sheet['A1'])
c=sheet['B1']
print(c.value)
print(c.coordinate)
print(c.row)
print(c.column)
print(sheet.cell(row=1, column=2).value)
for i in range(1,8,2):
    print(i,str(sheet.cell(row=i, column=2).value))
print(sheet.max_row)
print(sheet.max_column)
